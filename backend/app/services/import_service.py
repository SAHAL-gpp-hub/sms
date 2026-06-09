from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.base_models import (
    AcademicYear,
    Class,
    ImportBatch,
    ImportBatchItem,
    ImportStatusEnum,
    Student,
    StudentStatusEnum,
)
from app.schemas.student import StudentCreate
from app.services import student_service
from app.services.yearend_service import normalize_class_name

MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
SUPPORTED_EXTENSIONS = {'.csv', '.xlsx'}

REQUIRED_COLUMNS = {
    'name_en',
    'name_gu',
    'dob',
    'gender',
    'class_name',
    'father_name',
    'contact',
    'admission_date',
}

HEADER_ALIASES = {
    'class': 'class_name',
    'class_standard': 'class_name',
    'standard': 'class_name',
    'class_division': 'division',
    'section': 'division',
    'academic_year': 'academic_year_label',
    'year': 'academic_year_label',
    'mobile': 'contact',
    'phone': 'contact',
    'guardian_name': 'father_name',
}

# ---------------------------------------------------------------------------
# Column definitions — single source of truth for template, sample, and the
# Field Guide sheet.
#
# Tuple layout: (csv_key, human_label, example_value, notes, is_required)
#
# Why `student_id` is absent:
#   It is an internal DB surrogate auto-generated on insert. Showing it in the
#   template implies users must supply it, which they never do for new imports.
#   Legacy/migration files that DO include a student_id column still work fine
#   because _build_preview reads it via raw_row.get('student_id', '') and
#   passes it as student_id_override; absent = None, which is handled safely.
# ---------------------------------------------------------------------------
TEMPLATE_COLUMNS: list[tuple[str, str, str, str, bool]] = [
    # csv_key               human_label            example_value          notes                               required
    ('gr_number',           'GR Number',            'GR2024001',           'Leave blank to auto-generate',     False),
    ('name_en',             'Name (English)',        'Aarav Patel',         'Full name in English',             True),
    ('name_gu',             'Name (Gujarati)',       'આરવ પટેલ',            'Full name in Gujarati',            True),
    ('dob',                 'Date of Birth',         '2013-05-12',          'YYYY-MM-DD  e.g. 2013-05-12',      True),
    ('gender',              'Gender',                'M',                   'M / F / Other',                    True),
    ('class_name',          'Class',                 '7',                   'e.g. 7 or 7th',                    True),
    ('division',            'Division',              'A',                   'A, B, C … (default A)',            False),
    ('father_name',         'Father Name',           'Rakesh Patel',        'Father or primary guardian name',  True),
    ('mother_name',         'Mother Name',           'Pooja Patel',         '',                                 False),
    ('contact',             'Contact Number',        '9876543210',          '10-digit mobile number',           True),
    ('student_email',       'Student Email',         'aarav@example.com',   '',                                 False),
    ('student_phone',       'Student Phone',         '9876543210',          'Defaults to Contact if blank',     False),
    ('guardian_email',      'Guardian Email',        'rakesh@example.com',  '',                                 False),
    ('guardian_phone',      'Guardian Phone',        '9876543210',          'Defaults to Contact if blank',     False),
    ('address',             'Address',               'Palanpur, Gujarat',   '',                                 False),
    ('category',            'Category',              'GEN',                 'GEN / OBC / SC / ST',              False),
    ('aadhar_last4',        'Aadhaar Last 4',        '1234',                'Last 4 digits only',               False),
    ('admission_date',      'Admission Date',        '2024-06-01',          'YYYY-MM-DD  e.g. 2024-06-01',      True),
    ('academic_year_label', 'Academic Year',         '2025-26',             'Defaults to current year if blank',False),
    ('roll_number',         'Roll Number',           '1',                   'Positive whole number',            False),
    ('previous_school',     'Previous School',       'Iqra Primary School', '',                                 False),
]

_CSV_KEYS = [col[0] for col in TEMPLATE_COLUMNS]
_EXAMPLES = [col[2] for col in TEMPLATE_COLUMNS]


# ---------------------------------------------------------------------------
# XLSX template / sample builder
# ---------------------------------------------------------------------------

_NAV  = '1B3A6B'   # dark navy   — title banner / key header
_BLUE = '2563EB'   # medium blue — field guide header
_REQ  = 'FEF3C7'   # amber tint  — required columns
_OPT  = 'F1F5F9'   # slate tint  — optional columns
_SAM  = 'ECFDF5'   # green tint  — sample row
_ALT  = 'F8FAFC'   # faint stripe for data rows


def _border(color: str = 'CBD5E1') -> Border:
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _build_xlsx(include_sample: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Import Data'
    ws.sheet_view.showGridLines = False

    ncols = len(TEMPLATE_COLUMNS)
    col_widths = {
        'gr_number': 14, 'name_en': 22, 'name_gu': 22, 'dob': 14,
        'gender': 9, 'class_name': 9, 'division': 9, 'father_name': 20,
        'mother_name': 20, 'contact': 15, 'student_email': 26,
        'student_phone': 15, 'guardian_email': 26, 'guardian_phone': 15,
        'address': 26, 'category': 11, 'aadhar_last4': 14,
        'admission_date': 15, 'academic_year_label': 15,
        'roll_number': 12, 'previous_school': 24,
    }

    # Row 1 – banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    banner = ws.cell(row=1, column=1,
        value='Student Import  ·  Fill data from row 4 onwards  ·  Do NOT rename column headers in row 2')
    banner.font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    banner.fill      = PatternFill('solid', fgColor=_NAV)
    banner.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    for ci, (key, label, example, notes, required) in enumerate(TEMPLATE_COLUMNS, 1):
        bg = _REQ if required else _OPT

        # Row 2 – machine key (the column name the importer reads)
        c2 = ws.cell(row=2, column=ci, value=key)
        c2.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c2.fill      = PatternFill('solid', fgColor=_NAV)
        c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c2.border    = _border('334155')

        # Row 3 – friendly label + hint
        c3 = ws.cell(row=3, column=ci, value=f'{label}\n{notes}' if notes else label)
        c3.font      = Font(name='Arial', bold=True, size=8,
                            color='92400E' if required else '1E3A5F')
        c3.fill      = PatternFill('solid', fgColor=bg)
        c3.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c3.border    = _border()

        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(key, 15)

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 38

    data_start = 4
    if include_sample:
        for ci, (key, label, example, notes, required) in enumerate(TEMPLATE_COLUMNS, 1):
            c4 = ws.cell(row=4, column=ci, value=example)
            c4.font      = Font(name='Arial', size=9, italic=True, color='065F46')
            c4.fill      = PatternFill('solid', fgColor=_SAM)
            c4.alignment = Alignment(horizontal='left', vertical='center')
            c4.border    = _border('6EE7B7')
        ws.row_dimensions[4].height = 18
        data_start = 5

    # 200 data entry rows with alternating shading
    for row in range(data_start, data_start + 200):
        fill_color = 'FFFFFF' if row % 2 == 0 else _ALT
        for ci in range(1, ncols + 1):
            c = ws.cell(row=row, column=ci)
            c.fill      = PatternFill('solid', fgColor=fill_color)
            c.font      = Font(name='Arial', size=9)
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border    = _border()
        ws.row_dimensions[row].height = 18

    # Legend note below the data area
    legend_row = data_start + 201
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=ncols)
    leg = ws.cell(row=legend_row, column=1,
        value='🟡 Amber = REQUIRED column   ·   ⬜ Grey = optional   ·   Dates must be YYYY-MM-DD')
    leg.font      = Font(name='Arial', italic=True, size=8, color='64748B')
    leg.alignment = Alignment(horizontal='center')

    ws.freeze_panes = f'A{data_start}'

    # ── Sheet 2: Field Guide ───────────────────────────────────────────────
    guide = wb.create_sheet('Field Guide')
    guide.sheet_view.showGridLines = False
    guide.column_dimensions['A'].width = 22
    guide.column_dimensions['B'].width = 20
    guide.column_dimensions['C'].width = 36
    guide.column_dimensions['D'].width = 12

    for ci, heading in enumerate(['Column Key', 'Label', 'Notes / Accepted Values', 'Required?'], 1):
        c = guide.cell(row=1, column=ci, value=heading)
        c.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor=_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _border('1D4ED8')
    guide.row_dimensions[1].height = 22

    for ri, (key, label, _ex, notes, required) in enumerate(TEMPLATE_COLUMNS, 2):
        row_bg = _REQ if required else 'FFFFFF'
        vals   = [key, label, notes or '—', '✓ Required' if required else 'Optional']
        colors = ['1B3A6B', '1B3A6B', '374151', '166534' if required else '64748B']
        bolds  = [True, False, False, False]
        for ci, (val, color, bold) in enumerate(zip(vals, colors, bolds), 1):
            c = guide.cell(row=ri, column=ci, value=val)
            c.font      = Font(name='Arial', size=9, color=color, bold=bold)
            c.fill      = PatternFill('solid', fgColor=row_bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border    = _border()
        guide.row_dimensions[ri].height = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_template_xlsx() -> bytes:
    """Blank import template XLSX — header rows only, no sample data."""
    return _build_xlsx(include_sample=False)


def get_sample_xlsx() -> bytes:
    """Sample XLSX — header rows + one filled example row + Field Guide sheet."""
    return _build_xlsx(include_sample=True)


# ---------------------------------------------------------------------------
# Import pipeline (unchanged)
# ---------------------------------------------------------------------------

@dataclass
class RowContext:
    row_number: int
    raw: dict[str, Any]
    normalized: dict[str, Any]
    issues: list[str]
    warnings: list[str]
    status: str
    action: str
    class_key: tuple[str, str, int] | None
    missing_class: bool
    student_id: str | None
    gr_number: str | None


class ImportValidationError(ValueError):
    pass


def _normalize_header(value: Any) -> str:
    text = str(value or '').strip().lower()
    for ch in (' ', '-', '/', '\\'):
        text = text.replace(ch, '_')
    while '__' in text:
        text = text.replace('__', '_')
    return HEADER_ALIASES.get(text.strip('_'), text.strip('_'))


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _read_rows(filename: str, file_bytes: bytes) -> tuple[str, list[str], list[dict[str, str]]]:
    ext = Path(filename or '').suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ImportValidationError('Unsupported file type. Upload CSV or XLSX files only.')
    if len(file_bytes) > MAX_IMPORT_FILE_SIZE:
        raise ImportValidationError('Import file is too large. Maximum supported size is 5 MB.')

    if ext == '.csv':
        text = file_bytes.decode('utf-8-sig')
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames:
            raise ImportValidationError('CSV file is missing a header row.')
        headers = [_normalize_header(f) for f in reader.fieldnames]
        rows: list[dict[str, str]] = []
        for raw_row in reader:
            row = {headers[i]: _normalize_cell(v) for i, v in enumerate(raw_row.values()) if i < len(headers)}
            if any(row.values()):
                rows.append(row)
        return 'csv', headers, rows

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not raw_headers:
        raise ImportValidationError('XLSX file is missing a header row.')
    headers = [_normalize_header(f) for f in raw_headers]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: _normalize_cell(v) for i, v in enumerate(values) if i < len(headers)}
        if any(row.values()):
            rows.append(row)
    return 'xlsx', headers, rows


def _parse_date(value: str, field_name: str) -> date:
    cleaned = (value or '').strip()
    if not cleaned:
        raise ValueError(f'{field_name} is required')
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'{field_name} must be a valid date (YYYY-MM-DD preferred)')


def _normalize_gender(value: str) -> str:
    mapping = {'m': 'M', 'male': 'M', 'f': 'F', 'female': 'F', 'other': 'Other', 'o': 'Other'}
    cleaned = (value or '').strip().lower()
    if cleaned not in mapping:
        raise ValueError('gender must be M, F, or Other')
    return mapping[cleaned]


def _maybe_int(value: str, field_name: str) -> int | None:
    cleaned = (value or '').strip()
    if not cleaned:
        return None
    try:
        parsed = int(cleaned)
    except ValueError as exc:
        raise ValueError(f'{field_name} must be a whole number') from exc
    if parsed <= 0:
        raise ValueError(f'{field_name} must be greater than 0')
    return parsed


def _current_year(db: Session) -> AcademicYear | None:
    return db.query(AcademicYear).filter(AcademicYear.is_current.is_(True)).first()


def _build_class_lookup(db: Session) -> dict[tuple[str, str, int], Class]:
    lookup: dict[tuple[str, str, int], Class] = {}
    for cls in db.query(Class).all():
        key = (normalize_class_name(cls.name), (cls.division or 'A').strip().upper(), cls.academic_year_id)
        lookup[key] = cls
    return lookup


def _build_preview(
    db: Session,
    *,
    filename: str,
    file_bytes: bytes,
    create_missing_classes: bool,
) -> dict[str, Any]:
    file_format, headers, rows = _read_rows(filename, file_bytes)
    missing_columns = sorted(REQUIRED_COLUMNS - set(headers))

    year_lookup   = {y.label: y for y in db.query(AcademicYear).all()}
    current_year  = _current_year(db)
    class_lookup  = _build_class_lookup(db)

    incoming_student_ids = [r.get('student_id', '').strip() for r in rows if r.get('student_id', '').strip()]
    incoming_gr_numbers  = [r.get('gr_number', '').strip()  for r in rows if r.get('gr_number', '').strip()]

    existing_student_ids = (
        {v for v, in db.query(Student.student_id).filter(Student.student_id.in_(incoming_student_ids)).all()}
        if incoming_student_ids else set()
    )
    existing_gr_numbers = (
        {v for v, in db.query(Student.gr_number).filter(Student.gr_number.in_(incoming_gr_numbers)).all() if v}
        if incoming_gr_numbers else set()
    )

    file_student_id_counts = Counter(incoming_student_ids)
    file_gr_counts         = Counter(incoming_gr_numbers)

    preview_rows: list[RowContext] = []
    missing_class_refs: set[tuple[str, str, str]] = set()

    for index, raw_row in enumerate(rows, start=2):
        issues:   list[str] = []
        warnings: list[str] = []
        normalized: dict[str, Any] = {}

        student_id    = raw_row.get('student_id', '').strip() or None
        gr_number     = raw_row.get('gr_number', '').strip() or None
        class_name    = normalize_class_name(raw_row.get('class_name', '').strip()) if raw_row.get('class_name') else ''
        division      = (raw_row.get('division') or 'A').strip().upper() or 'A'
        year_label    = (raw_row.get('academic_year_label') or '').strip()
        academic_year = year_lookup.get(year_label) if year_label else current_year

        if missing_columns:
            issues.extend(f'Missing required column: {col}' for col in missing_columns)
        if not academic_year:
            issues.append('No academic year found for this row and no current academic year is set')
        if not class_name:
            issues.append('class_name is required')

        class_key = class_obj = None
        if academic_year and class_name:
            class_key = (class_name, division, academic_year.id)
            class_obj = class_lookup.get(class_key)
            if not class_obj:
                if create_missing_classes:
                    missing_class_refs.add((class_name, division, academic_year.label))
                    warnings.append(f'Class {class_name}-{division} will be created in {academic_year.label}')
                else:
                    issues.append(f'Class {class_name}-{division} does not exist in {academic_year.label}')

        if student_id and file_student_id_counts[student_id] > 1:
            issues.append('Duplicate student_id found in uploaded file')
        if gr_number and file_gr_counts[gr_number] > 1:
            issues.append('Duplicate gr_number found in uploaded file')
        if student_id and student_id in existing_student_ids:
            issues.append('student_id already exists in the system')
        if gr_number and gr_number in existing_gr_numbers:
            issues.append('gr_number already exists in the system')

        for field, parser in (
            ('dob',            lambda v: _parse_date(v, 'dob')),
            ('admission_date', lambda v: _parse_date(v, 'admission_date')),
            ('gender',         _normalize_gender),
            ('roll_number',    lambda v: _maybe_int(v, 'roll_number')),
        ):
            try:
                normalized[field] = parser(raw_row.get(field, ''))
            except ValueError as exc:
                issues.append(str(exc))

        contact = ''.join(ch for ch in raw_row.get('contact', '').strip() if ch.isdigit())
        normalized.update({
            'name_en':          raw_row.get('name_en', '').strip(),
            'name_gu':          raw_row.get('name_gu', '').strip(),
            'father_name':      raw_row.get('father_name', '').strip(),
            'mother_name':      raw_row.get('mother_name', '').strip() or None,
            'contact':          contact,
            'student_email':    raw_row.get('student_email', '').strip() or None,
            'student_phone':    ''.join(ch for ch in raw_row.get('student_phone', '').strip() if ch.isdigit()) or None,
            'guardian_email':   raw_row.get('guardian_email', '').strip() or None,
            'guardian_phone':   ''.join(ch for ch in raw_row.get('guardian_phone', '').strip() if ch.isdigit()) or None,
            'address':          raw_row.get('address', '').strip() or None,
            'category':         raw_row.get('category', '').strip() or None,
            'aadhar_last4':     raw_row.get('aadhar_last4', '').strip() or None,
            'academic_year_id': academic_year.id if academic_year else 1,
            'class_id':         class_obj.id if class_obj else 1,
            'gr_number':        gr_number,
            'previous_school':  raw_row.get('previous_school', '').strip() or None,
        })

        if contact:
            if not normalized.get('student_phone'):
                normalized['student_phone'] = contact
            if not normalized.get('guardian_phone'):
                normalized['guardian_phone'] = contact

        for field in ('name_en', 'name_gu', 'father_name', 'contact'):
            if not normalized.get(field):
                issues.append(f'{field} is required')

        if not issues:
            try:
                model = StudentCreate(**normalized)
                normalized = model.model_dump()
                normalized['previous_school'] = raw_row.get('previous_school', '').strip() or None
            except Exception as exc:  # pragma: no cover
                issues.append(str(exc).split('\n')[0])

        if issues and any('already exists' in i or 'Duplicate' in i for i in issues):
            action = 'skip_duplicate'
        elif issues:
            action = 'fix_errors'
        elif class_obj is None and class_key is not None:
            action = 'create_class_and_student'
        else:
            action = 'create'

        preview_rows.append(RowContext(
            row_number=index,
            raw=raw_row,
            normalized=normalized,
            issues=issues,
            warnings=warnings,
            status='ready' if not issues else 'invalid',
            action=action,
            class_key=class_key,
            missing_class=class_obj is None and class_key is not None,
            student_id=student_id,
            gr_number=gr_number,
        ))

    summary = {
        'total_rows':         len(preview_rows),
        'ready_rows':         sum(1 for r in preview_rows if r.status == 'ready'),
        'invalid_rows':       sum(1 for r in preview_rows if r.status == 'invalid'),
        'duplicate_rows':     sum(1 for r in preview_rows if any('already exists' in i or 'Duplicate' in i for i in r.issues)),
        'missing_class_rows': sum(1 for r in preview_rows if r.missing_class),
        'classes_to_create':  [
            {'class_name': cn, 'division': div, 'academic_year_label': yl}
            for cn, div, yl in sorted(missing_class_refs)
        ],
    }

    return {
        'file_name':        filename,
        'file_format':      file_format,
        'headers':          headers,
        'missing_columns':  missing_columns,
        'summary':          summary,
        'rows': [
            {
                'row_number':          r.row_number,
                'student_id':          r.student_id,
                'gr_number':           r.gr_number,
                'name_en':             r.raw.get('name_en', ''),
                'class_name':          r.raw.get('class_name', ''),
                'division':            r.raw.get('division') or 'A',
                'academic_year_label': r.raw.get('academic_year_label') or (current_year.label if current_year else ''),
                'status':              r.status,
                'action':              r.action,
                'issues':              r.issues,
                'warnings':            r.warnings,
            }
            for r in preview_rows
        ],
        '_contexts': preview_rows,
    }


def preview_students_import(
    db: Session,
    *,
    filename: str,
    file_bytes: bytes,
    create_missing_classes: bool,
) -> dict[str, Any]:
    try:
        payload = _build_preview(db, filename=filename, file_bytes=file_bytes,
                                 create_missing_classes=create_missing_classes)
    except ImportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    payload.pop('_contexts', None)
    return payload


def list_student_import_batches(db: Session, limit: int = 10) -> dict[str, Any]:
    items  = db.query(ImportBatch).filter_by(entity_type='student').order_by(ImportBatch.created_at.desc()).limit(limit).all()
    totals = db.query(ImportBatch).filter_by(entity_type='student').all()
    return {
        'summary': {
            'total_batches':       len(totals),
            'imported_rows':       sum(b.imported_rows for b in totals),
            'skipped_rows':        sum(b.skipped_rows  for b in totals),
            'error_rows':          sum(b.error_rows    for b in totals),
            'rolled_back_batches': sum(1 for b in totals if b.status == ImportStatusEnum.rolled_back),
        },
        'items': [
            {
                'id':              batch.id,
                'file_name':       batch.file_name,
                'file_format':     batch.file_format,
                'status':          batch.status.value if hasattr(batch.status, 'value') else batch.status,
                'merge_mode':      batch.merge_mode,
                'total_rows':      batch.total_rows,
                'imported_rows':   batch.imported_rows,
                'skipped_rows':    batch.skipped_rows,
                'error_rows':      batch.error_rows,
                'created_at':      batch.created_at.isoformat() if batch.created_at else None,
                'rolled_back_at':  batch.rolled_back_at.isoformat() if batch.rolled_back_at else None,
                'summary':         batch.summary or {},
                'created_by_name': batch.created_by.name if getattr(batch, 'created_by', None) else None,
            }
            for batch in items
        ],
    }


def commit_students_import(
    db: Session,
    *,
    filename: str,
    file_bytes: bytes,
    create_missing_classes: bool,
    actor: Any,
) -> dict[str, Any]:
    try:
        preview = _build_preview(db, filename=filename, file_bytes=file_bytes,
                                 create_missing_classes=create_missing_classes)
    except ImportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    contexts: list[RowContext] = preview.pop('_contexts')
    ready_rows = [r for r in contexts if r.status == 'ready']
    if not ready_rows:
        raise HTTPException(status_code=422, detail='No valid rows are ready to import')

    batch = ImportBatch(
        entity_type='student',
        file_name=filename,
        file_format=preview['file_format'],
        merge_mode='skip_duplicates',
        status=ImportStatusEnum.completed,
        total_rows=preview['summary']['total_rows'],
        imported_rows=0,
        skipped_rows=preview['summary']['duplicate_rows'],
        error_rows=preview['summary']['invalid_rows'] - preview['summary']['duplicate_rows'],
        summary={**preview['summary'], 'create_missing_classes': create_missing_classes},
        created_by_user_id=actor.id,
    )
    db.add(batch)
    db.flush()

    class_lookup = _build_class_lookup(db)
    created_classes: list[dict[str, Any]] = []

    if create_missing_classes:
        for ref in preview['summary']['classes_to_create']:
            academic_year = db.query(AcademicYear).filter_by(label=ref['academic_year_label']).first()
            if not academic_year:
                continue
            key = (normalize_class_name(ref['class_name']), ref['division'].strip().upper(), academic_year.id)
            if key in class_lookup:
                continue
            cls = Class(
                name=normalize_class_name(ref['class_name']),
                division=ref['division'].strip().upper(),
                academic_year_id=academic_year.id,
            )
            db.add(cls)
            db.flush()
            class_lookup[key] = cls
            created_classes.append({
                'id': cls.id, 'name': cls.name,
                'division': cls.division, 'academic_year_label': academic_year.label,
            })
            db.add(ImportBatchItem(
                import_batch_id=batch.id, entity_type='class', entity_id=cls.id, action='created',
                payload={'name': cls.name, 'division': cls.division, 'academic_year_label': academic_year.label},
            ))

    created_students: list[dict[str, Any]] = []
    try:
        current_year = _current_year(db)
        for row in ready_rows:
            academic_year_label = row.raw.get('academic_year_label') or (current_year.label if current_year else '')
            if not row.class_key:
                raise HTTPException(status_code=422, detail=f'Row {row.row_number} is missing class information')
            class_obj = class_lookup.get(row.class_key)
            if not class_obj:
                raise HTTPException(status_code=422, detail=f'Class missing for row {row.row_number}')
            payload = {**row.normalized, 'class_id': class_obj.id, 'academic_year_id': row.class_key[2]}
            student = student_service.create_student_from_import(
                db, StudentCreate(**payload),
                student_id_override=row.student_id,
                previous_school=payload.get('previous_school'),
                auto_commit=False,
            )
            created_students.append({'id': student.id, 'student_id': student.student_id, 'name_en': student.name_en})
            db.add(ImportBatchItem(
                import_batch_id=batch.id, entity_type='student', entity_id=student.id, action='created',
                payload={
                    'student_id': student.student_id, 'gr_number': student.gr_number,
                    'name_en': student.name_en, 'class_name': class_obj.name,
                    'division': class_obj.division, 'academic_year_label': academic_year_label,
                },
            ))

        batch.imported_rows = len(created_students)
        batch.summary = {**(batch.summary or {}),
                         'created_classes': created_classes,
                         'created_students_preview': created_students[:10]}
        db.commit()
    except Exception:
        db.rollback()
        raise

    db.refresh(batch)
    return {
        'batch': {
            'id': batch.id,
            'status': batch.status.value if hasattr(batch.status, 'value') else batch.status,
            'file_name': batch.file_name,
            'imported_rows': batch.imported_rows,
            'skipped_rows': batch.skipped_rows,
            'error_rows': batch.error_rows,
            'summary': batch.summary,
        },
        'preview': preview,
    }


def rollback_student_import(db: Session, batch_id: int, actor: Any) -> dict[str, Any]:
    batch = db.query(ImportBatch).filter_by(id=batch_id, entity_type='student').first()
    if not batch:
        raise HTTPException(status_code=404, detail='Import batch not found')
    if batch.status == ImportStatusEnum.rolled_back:
        raise HTTPException(status_code=409, detail='Import batch has already been rolled back')

    student_items = db.query(ImportBatchItem).filter_by(
        import_batch_id=batch.id, entity_type='student', action='created').all()
    class_items = db.query(ImportBatchItem).filter_by(
        import_batch_id=batch.id, entity_type='class', action='created').all()

    deactivated = skipped = 0
    for item in student_items:
        student = db.query(Student).filter_by(id=item.entity_id).first()
        if not student:
            skipped += 1
            continue
        if student.status != StudentStatusEnum.Left:
            student.status = StudentStatusEnum.Left
            if not student.reason_for_leaving:
                student.reason_for_leaving = f'Rolled back import batch #{batch.id}'
            deactivated += 1
        else:
            skipped += 1

    batch.status = ImportStatusEnum.rolled_back
    batch.rolled_back_at = datetime.now(UTC)
    batch.rollback_summary = {
        'rolled_back_by_user_id': actor.id,
        'deactivated_students':   deactivated,
        'skipped_students':       skipped,
        'preserved_classes':      len(class_items),
    }
    db.commit()
    return {
        'batch_id':         batch.id,
        'status':           batch.status.value if hasattr(batch.status, 'value') else batch.status,
        'rollback_summary': batch.rollback_summary,
    }